import glob
import os
import csv

from openpyxl import load_workbook

# main file to write to
stacked_file = 'co-cannabis-sales.csv'

# common headers for CSV files
headers = [
    'month',
    'year',
    'county',
    'amount',
    'sales_type'
]


def extract_pot_data(xlsx_filename):
    '''
    Given a monthly marijuana sales report, extract the data
    and return a tuple with the CSV filename (YYYYMM.csv)
    and the data as a list.
    '''

    # get the bare filename, without directories
    filename = xlsx_filename.split('/')[-1]

    # the first bit of the filename is the year and month
    namesplit = filename.split('_')[0]
    year, month = namesplit[:4], namesplit[-2:]
    fname = f'{year}{month}.csv'

    # load the workbook with openpyxl
    wb = load_workbook(xlsx_filename)
    ws = wb.active

    # a list to hold the data
    out = []

    # a flag to manage data collection as we iterate over the rows
    go = False

    # iterate over the worksheet rows
    for x in ws.iter_rows():

        # check if there's a value in cell of the first column
        if x[0].value:

            # get a casefolded version of that value
            first = x[0].value.casefold()

            # if we're at the "Total" line, we're done
            if first.startswith('total'):
                go = False
                break

            # if we're at the row with MMJ sales, flip the `go` flag and start collecting  # noqa
            if first.startswith('medical marijuana sales'):
                go = True
                continue

            # if we're in collection mode, get the data
            if go:

                # list comprehension to get the values in this row
                row = [z.value for z in x]

                # ignore the header row and some garbage over-wrapped rows
                if not row[0].startswith(('Sales', 'County')):

                    # extract counties and totals for both rec and medical tables  # noqa
                    mc, mt, _, rc, rt, *rest = row

                    # make sure there are values, and double-check that we're not at the summary row  # noqa
                    if mc and mt and not mc.startswith('Total'):
                        # some cleanup on the "Sum of NR Counties" row
                        mc = mc.replace('Counties 4', 'Counties')

                        # replacing their bespoke "no value" value with None
                        if mt == 'NR':
                            mt = None

                        # add this to the outlist
                        out.append([month, year, mc, mt, 'medical'])

                    # repeat this process for recreational totals
                    if rc and rt and not rc.startswith('Total'):
                        rc = rc.replace('Counties 4', 'Counties')
                        if rt == 'NR':
                            rt = None
                        out.append([month, year, rc, rt, 'retail'])

    # marry up headers w/ data
    data = [dict(zip(headers, x)) for x in out]

    # return a tuple with filename and data
    return (fname, data)


# get a list of raw files
raw_files = glob.glob(os.path.join('raw-data', '*.xlsx'))

with open(stacked_file, 'w') as mainfile:
    mainwriter = csv.DictWriter(mainfile, fieldnames=headers)
    mainwriter.writeheader()

    # loop over the xlsx files we downloaded
    for f in raw_files:

        # call the `extract_pot_data` function on the sheet
        # and assign filename and data variables
        fname, data = extract_pot_data(f)

        # build path to the CSV we're creating
        dest = os.path.join('processed-data', fname)

        # open that file to write to
        with open(dest, 'w') as monthly_file:

            # create a csv writer object
            monthlywriter = csv.DictWriter(monthly_file, fieldnames=headers)

            # write the headers to monthly file
            monthlywriter.writeheader()

            # write the data to monthly file
            monthlywriter.writerows(data)

            # write the data to main file
            mainwriter.writerows(data)
